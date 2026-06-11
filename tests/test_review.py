import sqlite3
from datetime import UTC, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from nist_fingerprint_comparator.core.models import NistTransaction
from nist_fingerprint_comparator.core.review import (
    DISPLAY_HISTORY_COLUMNS,
    HISTORY_COLUMNS,
    DecisionHistoryStore,
    DecisionXlsxExporter,
    ReviewDecision,
    ReviewQueue,
    available_export_path,
    decision_record,
)


def _transaction(path: Path, control_number: str) -> NistTransaction:
    path.write_bytes(f"transaction-{control_number}".encode())
    return NistTransaction(
        source_path=path,
        transaction_metadata={"1.009": control_number},
        descriptive_metadata={"MN1": f"MN1-{control_number}"},
    )


def _decision(
    tmp_path: Path,
    suffix: str,
    timestamp: str,
    value: str = "MATCH",
) -> ReviewDecision:
    return ReviewDecision(
        decision=value,  # type: ignore[arg-type]
        candidate_number=1,
        candidate_total=1,
        file_a=_transaction(tmp_path / f"a-{suffix}.nist", f"A-{suffix}"),
        file_b=_transaction(tmp_path / f"b-{suffix}.nist", f"B-{suffix}"),
        timestamp_utc=timestamp,
        timestamp=timestamp,
        timezone="UTC",
    )


def test_review_queue_walks_one_reference_against_many_candidates(tmp_path: Path) -> None:
    file_a = _transaction(tmp_path / "a.nist", "A-001")
    candidates = [tmp_path / "b1.nist", tmp_path / "b2.nist", tmp_path / "b3.nist"]
    queue = ReviewQueue()
    queue.start(file_a, candidates)

    for index, path in enumerate(candidates, start=1):
        file_b = _transaction(path, f"B-{index:03}")
        queue.record("MATCH" if index == 2 else "NO_MATCH", file_b)

    assert queue.is_complete
    assert [decision.decision for decision in queue.decisions] == [
        "NO_MATCH",
        "MATCH",
        "NO_MATCH",
    ]


def test_review_queue_accepts_pass_decision(tmp_path: Path) -> None:
    file_a = _transaction(tmp_path / "a.nist", "A-001")
    file_b = _transaction(tmp_path / "b.nist", "B-001")
    queue = ReviewQueue()
    queue.start(file_a, [file_b.source_path])

    queue.record("PASS", file_b)

    assert queue.decisions[0].decision == "PASS"


def test_review_queue_replaces_one_decision_per_candidate(tmp_path: Path) -> None:
    file_a = _transaction(tmp_path / "a.nist", "A-001")
    file_b = _transaction(tmp_path / "b.nist", "B-001")
    queue = ReviewQueue()
    queue.start(file_a, [file_b.source_path])

    first, previous = queue.set_decision("MATCH", file_b)
    replacement, previous = queue.set_decision("NO_MATCH", file_b)

    assert previous is first
    assert queue.decisions == [replacement]
    assert queue.is_complete


def test_review_queue_keeps_file_a_when_manually_selected_as_candidate(tmp_path: Path) -> None:
    file_a = _transaction(tmp_path / "a.nist", "A-001")
    queue = ReviewQueue()

    queue.start(file_a, [file_a.source_path])

    assert queue.current_path == file_a.source_path
    assert queue.candidate_total == 1


def test_internal_history_accumulates_across_future_sessions(tmp_path: Path) -> None:
    database = tmp_path / "history.sqlite3"
    first = DecisionHistoryStore(database)
    first.append(_decision(tmp_path, "1", "2026-06-10T08:00:00+00:00"))

    reopened = DecisionHistoryStore(database)
    reopened.append(_decision(tmp_path, "2", "2026-06-10T09:00:00+00:00", "NO_MATCH"))

    assert reopened.count() == 2
    assert [row["decision"] for row in reopened.query()] == ["MATCH", "NO_MATCH"]


def test_internal_history_query_filters_utc_time_range(tmp_path: Path) -> None:
    store = DecisionHistoryStore(tmp_path / "history.sqlite3")
    store.append(_decision(tmp_path, "1", "2026-06-09T23:00:00+00:00"))
    store.append(_decision(tmp_path, "2", "2026-06-10T12:00:00+00:00"))
    store.append(_decision(tmp_path, "3", "2026-06-11T01:00:00+00:00"))

    rows = store.query(
        datetime(2026, 6, 10, 0, 0, tzinfo=UTC),
        datetime(2026, 6, 10, 23, 59, tzinfo=UTC),
    )

    assert [row["file_b_reference_number"] for row in rows] == ["MN1-B-2"]
    assert rows[0]["timestamp"] == "12:00 10-06-2026"
    assert "file_a_transaction_control_number" not in rows[0]
    assert rows[0]["file_a_reference_number"] == "MN1-A-2"
    assert rows[0]["file_b_reference_number"] == "MN1-B-2"


def test_internal_history_can_delete_exact_session_decision(tmp_path: Path) -> None:
    store = DecisionHistoryStore(tmp_path / "history.sqlite3")
    first = _decision(tmp_path, "1", "2026-06-10T08:00:00+00:00")
    second = _decision(tmp_path, "2", "2026-06-10T09:00:00+00:00", "NO_MATCH")
    store.append(first)
    store.append(second)

    store.delete(second)

    assert store.count() == 1
    assert store.query()[0]["file_b_reference_number"] == "MN1-B-1"
    assert second.history_id is None


def test_internal_history_atomically_replaces_active_decision(tmp_path: Path) -> None:
    store = DecisionHistoryStore(tmp_path / "history.sqlite3")
    first = _decision(tmp_path, "1", "2026-06-10T08:00:00+00:00")
    replacement = _decision(
        tmp_path,
        "1",
        "2026-06-10T09:00:00+00:00",
        "NO_MATCH",
    )
    store.append(first)

    store.replace(first, replacement)

    assert first.history_id is None
    assert replacement.history_id is not None
    assert store.count() == 1
    assert store.query()[0]["decision"] == "NO_MATCH"


def test_internal_history_can_delete_record_by_hidden_history_id(tmp_path: Path) -> None:
    store = DecisionHistoryStore(tmp_path / "history.sqlite3")
    first = _decision(tmp_path, "1", "2026-06-10T08:00:00+00:00")
    second = _decision(tmp_path, "2", "2026-06-10T09:00:00+00:00", "NO_MATCH")
    store.append(first)
    store.append(second)
    history_id = int(store.query()[0]["history_id"])

    store.delete_by_id(history_id)

    assert store.count() == 1
    assert store.query()[0]["decision"] == "NO_MATCH"


def test_internal_history_can_delete_all_records(tmp_path: Path) -> None:
    store = DecisionHistoryStore(tmp_path / "history.sqlite3")
    store.append(_decision(tmp_path, "1", "2026-06-10T08:00:00+00:00"))
    store.append(_decision(tmp_path, "2", "2026-06-10T09:00:00+00:00", "NO_MATCH"))

    deleted = store.clear()

    assert deleted == 2
    assert store.count() == 0
    assert store.query() == []


def test_pass_decision_cannot_be_written_to_internal_history(tmp_path: Path) -> None:
    store = DecisionHistoryStore(tmp_path / "history.sqlite3")

    with pytest.raises(ValueError, match="PASS is not saved"):
        store.append(_decision(tmp_path, "1", "2026-06-10T08:00:00+00:00", "PASS"))

    assert store.count() == 0


def test_existing_history_database_removes_sha_columns_and_pass_rows(tmp_path: Path) -> None:
    database = tmp_path / "history.sqlite3"
    with sqlite3.connect(database) as connection:
        connection.execute(
            """
            CREATE TABLE decisions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp_utc TEXT NOT NULL,
                decision TEXT NOT NULL CHECK(decision IN ('MATCH', 'NO_MATCH', 'PASS')),
                file_a_name TEXT NOT NULL,
                file_a_sha256 TEXT NOT NULL,
                file_a_transaction_control_number TEXT NOT NULL,
                file_b_name TEXT NOT NULL,
                file_b_sha256 TEXT NOT NULL,
                file_b_transaction_control_number TEXT NOT NULL
            )
            """
        )
        connection.execute(
            "INSERT INTO decisions VALUES "
            "(1, '2026-06-10T08:00:00+00:00', 'MATCH', 'a.nist', '', 'A', "
            "'b.nist', '', 'B')"
        )
        connection.execute(
            "INSERT INTO decisions VALUES "
            "(2, '2026-06-10T09:00:00+00:00', 'PASS', 'a.nist', '', 'A', "
            "'c.nist', '', 'C')"
        )

    store = DecisionHistoryStore(database)
    store.append(_decision(tmp_path, "3", "2026-06-10T10:00:00+00:00", "NO_MATCH"))

    with sqlite3.connect(database) as connection:
        columns = [row[1] for row in connection.execute("PRAGMA table_info(decisions)")]
        saved_timestamps = connection.execute(
            "SELECT timestamp FROM decisions ORDER BY id"
        ).fetchall()

    assert columns == ["id", *HISTORY_COLUMNS]
    assert saved_timestamps == [("08:00 10-06-2026",), ("10:00 10-06-2026",)]
    assert [row["decision"] for row in store.query()] == ["MATCH", "NO_MATCH"]
    assert store.query()[0]["file_a_reference_number"] == ""
    assert store.query()[0]["file_b_reference_number"] == ""
    assert all("sha256" not in column for column in HISTORY_COLUMNS)
    assert HISTORY_COLUMNS == [
        "timestamp_utc",
        "timestamp",
        "timezone",
        "decision",
        "file_a_name",
        "file_b_name",
        "file_a_reference_number",
        "file_b_reference_number",
    ]


def test_xlsx_export_contains_minimal_history_columns(tmp_path: Path) -> None:
    decision = _decision(tmp_path, "1", "2026-06-10T12:00:00+00:00", "NO_MATCH")
    output = tmp_path / "history.xlsx"

    DecisionXlsxExporter().export(output, [decision_record(decision)])

    workbook = load_workbook(output, read_only=True)
    rows = list(workbook["Comparison History"].iter_rows(values_only=True))
    assert len(rows) == 2
    assert len(rows[0]) == len(DISPLAY_HISTORY_COLUMNS)
    assert rows[1][2] == "NO_MATCH"
    assert rows[1][3] == "a-1.nist"
    assert rows[1][4] == "b-1.nist"
    assert rows[1][5] == "MN1-A-1"
    assert rows[1][6] == "MN1-B-1"
    assert rows[1][0] == "12:00 10-06-2026"
    assert all("SHA-256" not in str(header) for header in rows[0])
    assert "Reference Record Name" in rows[0]
    assert "Reference Record Reference Number (MN1)" in rows[0]
    assert "Comparison Record Name" in rows[0]
    assert "Comparison Record Reference Number (MN1)" in rows[0]
    assert "Comparison Record Transaction Control Number" not in rows[0]


def test_available_export_path_uses_first_free_numbered_name(tmp_path: Path) -> None:
    original = tmp_path / "session.xlsx"
    second = tmp_path / "session_2.xlsx"
    original.write_bytes(b"existing")
    second.write_bytes(b"existing")

    assert available_export_path(original) == tmp_path / "session_3.xlsx"


def test_queue_can_restore_candidate_after_persistence_failure(tmp_path: Path) -> None:
    file_a = _transaction(tmp_path / "a.nist", "A-001")
    file_b = _transaction(tmp_path / "b.nist", "B-001")
    queue = ReviewQueue()
    queue.start(file_a, [file_b.source_path])

    queue.record("MATCH", file_b)
    rolled_back = queue.rollback_last()

    assert queue.current_path == file_b.source_path
    assert queue.decisions == []
    assert rolled_back is not None
    assert rolled_back.decision == "MATCH"
