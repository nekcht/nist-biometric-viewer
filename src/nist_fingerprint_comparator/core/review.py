"""One-to-many review queue, internal history, and XLSX export."""

from __future__ import annotations

import os
import sqlite3
import tempfile
from contextlib import closing
from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from PySide6.QtCore import QDateTime, Qt, QTimeZone

from .models import NistTransaction

ReviewDecisionValue = Literal["MATCH", "NO_MATCH", "PASS"]
HISTORY_ID_KEY = "history_id"

HISTORY_COLUMNS = [
    "timestamp_utc",
    "timestamp",
    "timezone",
    "decision",
    "file_a_name",
    "file_b_name",
    "file_a_reference_number",
    "file_b_reference_number",
]

DISPLAY_HISTORY_COLUMNS = [column for column in HISTORY_COLUMNS if column != "timestamp_utc"]

EXPORT_HEADERS = {
    "timestamp_utc": "Timestamp UTC (Internal)",
    "timestamp": "Timestamp",
    "timezone": "Time Zone",
    "decision": "Decision",
    "file_a_name": "Reference Record Name",
    "file_b_name": "Comparison Record Name",
    "file_a_reference_number": "Reference Record Reference Number (MN1)",
    "file_b_reference_number": "Comparison Record Reference Number (MN1)",
}


@dataclass(slots=True)
class ReviewDecision:
    decision: ReviewDecisionValue
    candidate_number: int
    candidate_total: int
    file_a: NistTransaction
    file_b: NistTransaction
    timestamp_utc: str = field(default_factory=lambda: datetime.now(UTC).isoformat())
    timestamp: str = field(default_factory=lambda: datetime.now(UTC).isoformat())
    timezone: str = "UTC"
    history_id: int | None = None


@dataclass(slots=True)
class ReviewQueue:
    file_a: NistTransaction | None = None
    candidate_paths: list[Path] = field(default_factory=list)
    current_index: int = -1
    decisions: list[ReviewDecision] = field(default_factory=list)

    def start(self, file_a: NistTransaction, candidate_paths: list[Path]) -> None:
        self.file_a = file_a
        self.candidate_paths = list(dict.fromkeys(candidate_paths))
        self.current_index = 0 if self.candidate_paths else -1
        self.decisions.clear()

    @property
    def current_path(self) -> Path | None:
        if 0 <= self.current_index < len(self.candidate_paths):
            return self.candidate_paths[self.current_index]
        return None

    @property
    def candidate_number(self) -> int:
        if self.current_path is not None:
            return self.current_index + 1
        return len(self.candidate_paths)

    @property
    def candidate_total(self) -> int:
        return len(self.candidate_paths)

    @property
    def is_complete(self) -> bool:
        return bool(self.candidate_paths) and len(self.decisions) == len(
            self.candidate_paths
        )

    def set_current_index(self, index: int) -> None:
        if not 0 <= index < len(self.candidate_paths):
            raise IndexError("Comparison Record index is out of range.")
        self.current_index = index

    def decision_for_index(self, index: int) -> ReviewDecision | None:
        candidate_number = index + 1
        return next(
            (
                decision
                for decision in self.decisions
                if decision.candidate_number == candidate_number
            ),
            None,
        )

    def next_undecided_index(self, after_index: int) -> int | None:
        decided = {decision.candidate_number - 1 for decision in self.decisions}
        ordered_indexes = [
            *range(after_index + 1, len(self.candidate_paths)),
            *range(0, after_index),
        ]
        return next((index for index in ordered_indexes if index not in decided), None)

    def set_decision(
        self,
        decision: ReviewDecisionValue,
        file_b: NistTransaction,
    ) -> tuple[ReviewDecision, ReviewDecision | None]:
        if self.file_a is None or self.current_path is None:
            raise ValueError("No Comparison Record selected.")
        previous = self.decision_for_index(self.current_index)
        review_decision = ReviewDecision(
            decision=decision,
            candidate_number=self.current_index + 1,
            candidate_total=len(self.candidate_paths),
            file_a=self.file_a,
            file_b=file_b,
        )
        if previous is not None:
            self.decisions.remove(previous)
        self.decisions.append(review_decision)
        self.decisions.sort(key=lambda item: item.candidate_number)
        return review_decision, previous

    def restore_decision(
        self,
        candidate_index: int,
        previous: ReviewDecision | None,
    ) -> None:
        current = self.decision_for_index(candidate_index)
        if current is not None:
            self.decisions.remove(current)
        if previous is not None:
            self.decisions.append(previous)
            self.decisions.sort(key=lambda item: item.candidate_number)

    def record(
        self,
        decision: ReviewDecisionValue,
        file_b: NistTransaction,
    ) -> ReviewDecision:
        review_decision, _ = self.set_decision(decision, file_b)
        self.current_index += 1
        return review_decision

    def rollback_last(self) -> ReviewDecision | None:
        if not self.decisions:
            return None
        decision = self.decisions.pop()
        self.current_index -= 1
        return decision


class DecisionHistoryStore:
    """Persistent SQLite history for committed MATCH and NO_MATCH decisions."""

    def __init__(self, path: Path) -> None:
        self.path = path
        self._initialize()

    def append(self, decision: ReviewDecision) -> int:
        if decision.decision == "PASS":
            raise ValueError("PASS is not saved to History.")
        row = decision_record(decision)
        placeholders = ", ".join("?" for _ in HISTORY_COLUMNS)
        columns = ", ".join(HISTORY_COLUMNS)
        with closing(self._connect()) as connection, connection:
            cursor = connection.execute(
                f"INSERT INTO decisions ({columns}) VALUES ({placeholders})",
                [row[column] for column in HISTORY_COLUMNS],
            )
            history_id = int(cursor.lastrowid)
        decision.history_id = history_id
        return history_id

    def delete(self, decision: ReviewDecision) -> None:
        if decision.history_id is None:
            raise ValueError("Decision is not in History.")
        self.delete_by_id(decision.history_id)
        decision.history_id = None

    def replace(
        self,
        previous: ReviewDecision | None,
        decision: ReviewDecision,
    ) -> int | None:
        """Atomically replace one active-session history record."""
        row = decision_record(decision)
        placeholders = ", ".join("?" for _ in HISTORY_COLUMNS)
        columns = ", ".join(HISTORY_COLUMNS)
        with closing(self._connect()) as connection, connection:
            if previous is not None and previous.history_id is not None:
                cursor = connection.execute(
                    "DELETE FROM decisions WHERE id = ?",
                    (previous.history_id,),
                )
                if cursor.rowcount != 1:
                    raise ValueError("History record not found.")
            history_id = None
            if decision.decision != "PASS":
                cursor = connection.execute(
                    f"INSERT INTO decisions ({columns}) VALUES ({placeholders})",
                    [row[column] for column in HISTORY_COLUMNS],
                )
                history_id = int(cursor.lastrowid)
        if previous is not None:
            previous.history_id = None
        decision.history_id = history_id
        return history_id

    def delete_by_id(self, history_id: int) -> None:
        with closing(self._connect()) as connection, connection:
            cursor = connection.execute("DELETE FROM decisions WHERE id = ?", (history_id,))
            if cursor.rowcount != 1:
                raise ValueError("History record not found.")

    def clear(self) -> int:
        """Delete every persisted decision and return the number removed."""
        with closing(self._connect()) as connection, connection:
            cursor = connection.execute("DELETE FROM decisions")
            return max(cursor.rowcount, 0)

    def query(
        self,
        start_utc: datetime | None = None,
        end_utc: datetime | None = None,
        *,
        limit: int | None = None,
        offset: int = 0,
    ) -> list[dict[str, str]]:
        clauses: list[str] = []
        parameters: list[str | int] = []
        if start_utc is not None:
            clauses.append("timestamp_utc >= ?")
            parameters.append(_utc_iso(start_utc))
        if end_utc is not None:
            clauses.append("timestamp_utc <= ?")
            parameters.append(_utc_iso(end_utc))
        where = f" WHERE {' AND '.join(clauses)}" if clauses else ""
        pagination = ""
        if limit is not None:
            if limit <= 0 or offset < 0:
                raise ValueError("History page size and offset must be positive.")
            pagination = " LIMIT ? OFFSET ?"
            parameters.extend((limit, offset))
        with closing(self._connect()) as connection:
            rows = connection.execute(
                f"SELECT id, {', '.join(HISTORY_COLUMNS)} FROM decisions"
                f"{where} ORDER BY timestamp_utc DESC, id DESC{pagination}",
                parameters,
            ).fetchall()
        return [_history_row(row) for row in rows]

    def count(self) -> int:
        with closing(self._connect()) as connection:
            return int(connection.execute("SELECT COUNT(*) FROM decisions").fetchone()[0])

    def _initialize(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        with closing(self._connect()) as connection, connection:
            schema = connection.execute(
                "SELECT sql FROM sqlite_master WHERE type = 'table' AND name = 'decisions'"
            ).fetchone()
            if schema is not None and self._history_schema_requires_migration(
                connection,
                schema[0],
            ):
                self._migrate_history_schema(connection)
            connection.execute(_CREATE_HISTORY_TABLE_SQL)
            self._normalize_history_timestamps(connection)

    @staticmethod
    def _history_schema_requires_migration(
        connection: sqlite3.Connection,
        schema_sql: str,
    ) -> bool:
        columns = [
            row[1] for row in connection.execute("PRAGMA table_info(decisions)").fetchall()
        ]
        return columns != ["id", *HISTORY_COLUMNS] or "'PASS'" in schema_sql

    @staticmethod
    def _migrate_history_schema(connection: sqlite3.Connection) -> None:
        existing_columns = {
            row[1] for row in connection.execute("PRAGMA table_info(decisions)").fetchall()
        }
        connection.execute("ALTER TABLE decisions RENAME TO decisions_before_migration")
        connection.execute(_CREATE_HISTORY_TABLE_SQL)
        selections = []
        for column in HISTORY_COLUMNS:
            if column in existing_columns:
                selections.append(column)
            elif column == "timestamp":
                selections.append("timestamp_utc")
            elif column == "timezone":
                selections.append("'UTC'")
            elif column.endswith("_reference_number"):
                selections.append("''")
            else:
                raise sqlite3.DatabaseError(
                    f"History database is missing required column: {column}"
                )
        connection.execute(
            f"INSERT INTO decisions (id, {', '.join(HISTORY_COLUMNS)}) "
            f"SELECT id, {', '.join(selections)} FROM decisions_before_migration "
            "WHERE decision IN ('MATCH', 'NO_MATCH')"
        )
        connection.execute("DROP TABLE decisions_before_migration")

    @staticmethod
    def _normalize_history_timestamps(connection: sqlite3.Connection) -> None:
        rows = connection.execute(
            "SELECT id, timestamp_utc, timestamp, timezone FROM decisions"
        ).fetchall()
        updates = [
            (formatted, history_id)
            for history_id, timestamp_utc, timestamp, timezone in rows
            if (formatted := _formatted_timestamp(timestamp_utc, timezone)) != timestamp
        ]
        connection.executemany(
            "UPDATE decisions SET timestamp = ? WHERE id = ?",
            updates,
        )

    def _connect(self) -> sqlite3.Connection:
        return sqlite3.connect(self.path, timeout=5.0)


class DecisionXlsxExporter:
    """Export selected internal-history rows to a readable XLSX workbook."""

    def export(self, path: Path, rows: list[dict[str, str]]) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Comparison History"
        sheet.append([EXPORT_HEADERS[column] for column in DISPLAY_HISTORY_COLUMNS])
        for row in rows:
            sheet.append([row[column] for column in DISPLAY_HISTORY_COLUMNS])

        header_fill = PatternFill("solid", fgColor="D9E3EC")
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for index, column in enumerate(DISPLAY_HISTORY_COLUMNS, start=1):
            values = [EXPORT_HEADERS[column], *(str(row[column]) for row in rows)]
            sheet.column_dimensions[get_column_letter(index)].width = min(
                max(len(value) for value in values) + 2,
                72,
            )
        temporary_path: Path | None = None
        try:
            with tempfile.NamedTemporaryFile(
                prefix=f".{path.stem}-",
                suffix=".tmp",
                dir=path.parent,
                delete=False,
            ) as temporary_file:
                temporary_path = Path(temporary_file.name)
            workbook.save(temporary_path)
            os.replace(temporary_path, path)
            temporary_path = None
        finally:
            workbook.close()
            if temporary_path is not None:
                temporary_path.unlink(missing_ok=True)


def available_export_path(path: Path) -> Path:
    """Return the first numbered alternative that does not already exist."""
    if not path.exists():
        return path
    index = 2
    while True:
        candidate = path.with_name(f"{path.stem}_{index}{path.suffix}")
        if not candidate.exists():
            return candidate
        index += 1


def decision_record(decision: ReviewDecision) -> dict[str, str]:
    row = {
        "timestamp_utc": decision.timestamp_utc,
        "timestamp": _formatted_timestamp(decision.timestamp_utc, decision.timezone),
        "timezone": decision.timezone,
        "decision": decision.decision,
    }
    row.update(_transaction_fields("file_a", decision.file_a))
    row.update(_transaction_fields("file_b", decision.file_b))
    return row


def _transaction_fields(prefix: str, transaction: NistTransaction) -> dict[str, str]:
    fields = {
        f"{prefix}_name": transaction.source_path.name,
        f"{prefix}_reference_number": transaction.reference_number,
    }
    return fields


def _utc_iso(value: datetime) -> str:
    if value.tzinfo is None:
        value = value.replace(tzinfo=UTC)
    return value.astimezone(UTC).isoformat()


def _history_row(row: tuple) -> dict[str, str]:
    result = {
        HISTORY_ID_KEY: str(row[0]),
        **dict(zip(HISTORY_COLUMNS, row[1:], strict=True)),
    }
    result["timestamp"] = _formatted_timestamp(
        result["timestamp_utc"],
        result["timezone"],
    )
    return result


def _formatted_timestamp(timestamp_utc: str, timezone_id: str) -> str:
    timestamp = QDateTime.fromString(timestamp_utc, Qt.DateFormat.ISODateWithMs)
    if not timestamp.isValid():
        timestamp = QDateTime.fromString(timestamp_utc, Qt.DateFormat.ISODate)
    timezone = QTimeZone(timezone_id.encode())
    if not timezone.isValid():
        timezone = QTimeZone.utc()
    if not timestamp.isValid():
        return timestamp_utc
    return timestamp.toTimeZone(timezone).toString("HH:mm dd-MM-yyyy")


_CREATE_HISTORY_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS decisions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    timestamp_utc TEXT NOT NULL,
    timestamp TEXT NOT NULL,
    timezone TEXT NOT NULL,
    decision TEXT NOT NULL CHECK(decision IN ('MATCH', 'NO_MATCH')),
    file_a_name TEXT NOT NULL,
    file_b_name TEXT NOT NULL,
    file_a_reference_number TEXT NOT NULL,
    file_b_reference_number TEXT NOT NULL
)
"""
